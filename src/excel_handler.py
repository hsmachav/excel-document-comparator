import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, List, Tuple


class ExcelHandler:
    """Handles reading and writing Excel files."""

    @staticmethod
    def read_excel(file_path: str) -> pd.DataFrame:
        """
        Read Excel file and return DataFrame.
        Dynamically finds Name, Revision, and File name columns.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            DataFrame with columns: Document Name, Revision, File name
        """
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            
            # Strip whitespace from all string columns
            # Use apply instead of applymap (applymap deprecated in pandas 2.1.0+)
            df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x) if col.dtype == 'object' else col)
            
            # Find columns by matching keywords (case-insensitive)
            column_mapping = {}
            df_columns_lower = [col.lower() for col in df.columns]
            
            # Find Document Name column - look for "Name" (not "File Name" or "Native files count")
            for idx, col in enumerate(df_columns_lower):
                col_lower = col.strip()
                # Match "Name" but exclude "File Name", "Native files count"
                if col_lower == 'name':
                    column_mapping['Document Name'] = df.columns[idx]
                    break
            
            # Find Revision column
            for idx, col in enumerate(df_columns_lower):
                col_lower = col.strip()
                if col_lower == 'revision':
                    column_mapping['Revision'] = df.columns[idx]
                    break
            
            # Find File name column - case insensitive, look for "File name" or "File Name"
            for idx, col in enumerate(df_columns_lower):
                col_lower = col.strip()
                if col_lower == 'file name' or col_lower == 'filename':
                    column_mapping['File name'] = df.columns[idx]
                    break
            
            # Validate that all required columns were found
            if len(column_mapping) < 3:
                missing = []
                if 'Document Name' not in column_mapping:
                    missing.append('Name')
                if 'Revision' not in column_mapping:
                    missing.append('Revision')
                if 'File name' not in column_mapping:
                    missing.append('File name or File Name')
                
                raise Exception(
                    f"Could not find required columns: {', '.join(missing)}. "
                    f"Available columns: {list(df.columns)}"
                )
            
            # Select only the required columns
            df = df[list(column_mapping.values())]
            df.columns = list(column_mapping.keys())
            
            # Remove any rows with NaN in key columns
            df = df.dropna(subset=['Document Name', 'Revision', 'File name'])
            
            return df
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")

    @staticmethod
    def write_output_excel(output_file: str, changes: List[Dict], summary: Dict) -> None:
        """
        Write changes to output Excel file with formatting.
        
        Args:
            output_file: Path to output Excel file
            changes: List of change dictionaries
            summary: Summary statistics
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create Changes sheet
        if changes:
            ExcelHandler._create_changes_sheet(wb, changes)

        # Create Summary sheet
        ExcelHandler._create_summary_sheet(wb, summary)

        wb.save(output_file)

    @staticmethod
    def _create_changes_sheet(wb: Workbook, changes: List[Dict]) -> None:
        """
        Create the changes sheet with detailed information.
        
        Args:
            wb: Workbook object
            changes: List of change dictionaries
        """
        ws = wb.create_sheet("Changes", 0)

        # Define headers
        headers = [
            "Document Name",
            "Old Revision",
            "New Revision",
            "Old Files",
            "New Files",
            "Change Type",
            "Remarks"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write data
        for row, change in enumerate(changes, 2):
            ws.cell(row=row, column=1).value = change['document_name']
            ws.cell(row=row, column=2).value = change['old_revision']
            ws.cell(row=row, column=3).value = change['new_revision']
            ws.cell(row=row, column=4).value = change['old_files']
            ws.cell(row=row, column=5).value = change['new_files']
            ws.cell(row=row, column=6).value = change['change_type']
            ws.cell(row=row, column=7).value = change['remarks']

            # Apply alignment and wrapping
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 40

    @staticmethod
    def _create_summary_sheet(wb: Workbook, summary: Dict) -> None:
        """
        Create the summary sheet with statistics.
        
        Args:
            wb: Workbook object
            summary: Summary statistics
        """
        ws = wb.create_sheet("Summary", 1)

        # Title
        ws['A1'] = "Document Comparison Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Summary data
        row = 3
        summary_data = [
            ("Total Unique Documents Changed", summary['total_changed']),
            ("Total Changes Detected", summary['total_changes']),
            ("Revision Changed Only", summary['revision_only']),
            ("Files Changed Only", summary['files_only']),
            ("Both Revision and Files Changed", summary['both_changed'])
        ]

        for label, value in summary_data:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
